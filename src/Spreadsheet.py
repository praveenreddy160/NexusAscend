import openpyxl

# Load the workbook and select the sheet
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Initialize dictionaries to hold data
products_per_supplier = {}
total_value_per_supplier = {}
product_under_10 = {}

# Process each product starting from row 2
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value  # Supplier name
    inventory = product_list.cell(product_row, 2).value  # Inventory count
    price = product_list.cell(product_row, 3).value  # Product price
    product_num = product_list.cell(product_row, 1).value  # Product number
    inventory_price = product_list.cell(product_row, 5)  # Cell for inventory total price

    # Print supplier name
    print(f"Supplier Name: {supplier_name}")

    # Counting products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        print(f"Existing Supplier - {supplier_name}: Current Count = {current_num_products}")
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1
        print(f"New Supplier - {supplier_name}: Count = 1")

    # Calculation of total value per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        print(f"Current total value for {supplier_name}: {current_total_value}")
        total_value_per_supplier[supplier_name] = current_total_value + (inventory * price)
        print(total_value_per_supplier)
    else:
        total_value_per_supplier[supplier_name] = price * inventory
        print(f"Initial total value for {supplier_name}: {total_value_per_supplier[supplier_name]}")

    # Logic if inventory is less than 10
    if inventory < 10:
        product_under_10[int(product_num)] = inventory

    # Logic for total inventory price
    inventory_price.value = inventory * price

# Print the final results
print("Final Products Per Supplier:")
print(products_per_supplier)
print("Total Value Per Supplier:")
print(total_value_per_supplier)
print("Products Under 10 in Inventory:")
print(product_under_10)

# Save the updated workbook
inv_file.save("inventory_with_total_value.xlsx")